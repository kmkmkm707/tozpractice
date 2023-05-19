import pykeepass
from openpyxl import Workbook
from openpyxl.styles import Font

def get_group_entries(group):
    entries = []
    for entry in group.entries:
        entries.append(entry)
    for subgroup in group.subgroups:
        entries += get_group_entries(subgroup)
    return entries

def get_group_name(group, delimiter='\\'):
    if group:
        parent_name = get_group_name(group.parentgroup)
        if parent_name:
            return parent_name + delimiter + group.name
        else:
            return group.name
    else:
        return ''

db = pykeepass.PyKeePass('D:\keepass\DataBaseKeepass\DB.kdbx', password='O09qcqXsLKctQWERT=1234FDSA')
target_group_name = 'Учащиеся'
target_group = db.find_groups(name=target_group_name, first=True)

if target_group:
    workbook = Workbook()
    sheet = workbook.active
    sheet['A1'] = 'Группа'
    sheet['B1'] = 'Ф.И.О'
    sheet['C1'] = 'URL-ссылки'
    sheet['D1'] = 'Логин'
    sheet['E1'] = 'Пароль'
    sheet['F1'] = 'Заметки'
    entries = get_group_entries(target_group)

    for i, entry in enumerate(entries):
        group_name = get_group_name(entry.group)
        title = entry.title
        url = entry.url
        if url == None:
            url = '-'
        username = entry.username
        password = entry.password
        notes = entry.notes
        if notes == None:
            notes = '-'
        row = i + 2
        sheet.cell(row=row, column=1, value=group_name)
        sheet.cell(row=row, column=2, value=title).font = Font(bold=True)
        sheet.cell(row=row, column=3, value=url)
        sheet.cell(row=row, column=4, value=username).font = Font(bold=True)
        sheet.cell(row=row, column=5, value=password).font = Font(bold=True)
        sheet.cell(row=row, column=6, value=notes)
    workbook.save('D:\keepass\DataBaseKeepass\output.xlsx')
else:
    print(f"Нет группы с названием - {target_group_name}")
